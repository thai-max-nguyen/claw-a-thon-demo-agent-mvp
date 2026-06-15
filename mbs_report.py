"""MBS daily-report pipeline (Chị Nga / Biz ZLP workflow).

Flow:
  1. Load daily metrics (CSV — swap for the real MBS data source / SQL).
  2. Compute KPIs + day-over-day deltas + top movers.
  3. Write/update an Excel report (mbs_report.xlsx).
  4. Generate an executive narrative via GreenNode MaaS.
  5. Deliver: --telegram (post summary) and/or --confluence (paste to the
     "MVP Output MBS" page). Default = print the report only.

Then the operator can reply on Telegram to request an ACTION (push noti, drill-down),
handled by telegram_bot.py /action.

Env: LLM_BASE_URL, LLM_API_KEY, LLM_MODEL (report narrative uses MODEL_REPORT
or LLM_MODEL); TELEGRAM_BOT_TOKEN, TELEGRAM_CHAT_ID; Confluence token at
~/.config/confluence-token. Data source: MBS_DATA (default mbs/sample_data.csv).
"""
import csv
import os
import sys
import datetime as dt
from collections import defaultdict

DATA = os.getenv("MBS_DATA", os.path.join(os.path.dirname(__file__), "mbs", "sample_data.csv"))
XLSX = os.getenv("MBS_XLSX", os.path.join(os.path.dirname(__file__), "mbs_report.xlsx"))
LLM_BASE_URL = os.getenv("LLM_BASE_URL", "").rstrip("/")
LLM_API_KEY = os.getenv("LLM_API_KEY", "").strip()
MODEL_REPORT = os.getenv("MODEL_REPORT", "").strip() or os.getenv("LLM_MODEL", "google/gemma-4-31b-it")
CONFLUENCE_PAGE_ID = os.getenv("MBS_CONFLUENCE_PAGE_ID", "335581153")  # "MVP Output MBS"


def load_rows():
    with open(DATA) as f:
        return list(csv.DictReader(f))


def compute(rows):
    by_day = defaultdict(list)
    for r in rows:
        by_day[r["date"]].append(r)
    days = sorted(by_day)
    if not days:
        raise SystemExit("no data")
    today, prev = days[-1], (days[-2] if len(days) > 1 else None)

    def agg(day):
        rs = by_day[day]
        tx = sum(int(r["transactions"]) for r in rs)
        rev = sum(int(r["revenue_vnd"]) for r in rs)
        sr = sum(float(r["success_rate"]) * int(r["transactions"]) for r in rs) / max(tx, 1)
        return {"tx": tx, "rev": rev, "sr": sr, "rows": rs}

    cur = agg(today)
    pre = agg(prev) if prev else None
    movers = []
    if pre:
        pmap = {r["product"]: int(r["transactions"]) for r in pre["rows"]}
        for r in cur["rows"]:
            p, t = r["product"], int(r["transactions"])
            base = pmap.get(p, t)
            movers.append((p, t, (t - base) / base * 100 if base else 0))
        movers.sort(key=lambda m: abs(m[2]), reverse=True)
    return {"today": today, "prev": prev, "cur": cur, "pre": pre, "movers": movers}


def _pct(cur, pre):
    return (cur - pre) / pre * 100 if pre else 0.0


def write_excel(k):
    import openpyxl
    cur, pre = k["cur"], k["pre"]
    try:
        wb = openpyxl.load_workbook(XLSX)
    except Exception:
        wb = openpyxl.Workbook()
        wb.active.title = "Report"
    ws = wb["Report"] if "Report" in wb.sheetnames else wb.active
    ws.delete_rows(1, ws.max_row or 1)
    ws.append([f"MBS Daily Report — {k['today']}"])
    ws.append(["Metric", "Today", "Prev", "Δ %"])
    ws.append(["Transactions", cur["tx"], pre["tx"] if pre else "", round(_pct(cur["tx"], pre["tx"]) if pre else 0, 2)])
    ws.append(["Revenue (VND)", cur["rev"], pre["rev"] if pre else "", round(_pct(cur["rev"], pre["rev"]) if pre else 0, 2)])
    ws.append(["Success rate", round(cur["sr"], 4), round(pre["sr"], 4) if pre else "", round((cur["sr"] - pre["sr"]) * 100 if pre else 0, 2)])
    ws.append([])
    ws.append(["Product", "Transactions", "Revenue (VND)", "Success rate"])
    for r in cur["rows"]:
        ws.append([r["product"], int(r["transactions"]), int(r["revenue_vnd"]), float(r["success_rate"])])
    wb.save(XLSX)
    return XLSX


def narrative(k):
    cur, pre = k["cur"], k["pre"]
    facts = (f"Date {k['today']}. Transactions {cur['tx']:,} ({_pct(cur['tx'], pre['tx']):+.1f}% DoD). "
             f"Revenue {cur['rev']:,} VND ({_pct(cur['rev'], pre['rev']):+.1f}%). "
             f"Success rate {cur['sr']*100:.1f}%. "
             f"Top movers: " + "; ".join(f"{p} {d:+.1f}%" for p, _, d in k["movers"][:3]))
    if not (LLM_API_KEY and LLM_BASE_URL):
        return f"[stub narrative — set LLM_API_KEY] {facts}"
    try:
        from openai import OpenAI
        client = OpenAI(base_url=LLM_BASE_URL, api_key=LLM_API_KEY, timeout=60)
        r = client.chat.completions.create(
            model=MODEL_REPORT,
            messages=[{"role": "system", "content": "You are a fintech business analyst. Write a tight 4-6 sentence executive summary of the daily MBS metrics: headline, what moved, one risk/watch item. No preamble."},
                      {"role": "user", "content": facts}],
            max_tokens=400, temperature=0.5)
        return (r.choices[0].message.content or "").strip() or facts
    except Exception as e:
        return f"[narrative unavailable: {type(e).__name__}] {facts}"


def report_text(k):
    cur, pre = k["cur"], k["pre"]
    lines = [f"📊 MBS Daily Report — {k['today']}", "", narrative(k), "", "KPIs:",
             f"• Transactions: {cur['tx']:,} ({_pct(cur['tx'], pre['tx']):+.1f}% DoD)" if pre else f"• Transactions: {cur['tx']:,}",
             f"• Revenue: {cur['rev']:,} VND ({_pct(cur['rev'], pre['rev']):+.1f}%)" if pre else f"• Revenue: {cur['rev']:,} VND",
             f"• Success rate: {cur['sr']*100:.1f}%"]
    return "\n".join(lines)


def paste_confluence(k):
    """Append a dated report section to the MVP Output MBS page."""
    import urllib.request, base64, json
    auth = open(os.path.expanduser("~/.config/confluence-token")).read().strip()
    h = {"Authorization": "Basic " + base64.b64encode(auth.encode()).decode(), "Content-Type": "application/json"}
    base = "https://confluence.zalopay.vn/rest/api/content/" + CONFLUENCE_PAGE_ID
    req = urllib.request.Request(base + "?expand=body.storage,version", headers=h)
    d = json.load(urllib.request.urlopen(req, timeout=20))
    ver = d["version"]["number"]
    body = d["body"]["storage"]["value"]
    cur, pre = k["cur"], k["pre"]
    rows = "".join(f"<tr><td><p>{r['product']}</p></td><td><p>{int(r['transactions']):,}</p></td>"
                   f"<td><p>{int(r['revenue_vnd']):,}</p></td><td><p>{float(r['success_rate'])*100:.1f}%</p></td></tr>"
                   for r in cur["rows"])
    nar = narrative(k).replace("&", "&amp;").replace("<", "&lt;")
    section = (f"<h2>{k['today']} — MBS Daily Report</h2><ac:structured-macro ac:name=\"info\" ac:schema-version=\"1\" "
               f"ac:macro-id=\"{__import__('uuid').uuid4()}\"><ac:rich-text-body><p>{nar}</p></ac:rich-text-body></ac:structured-macro>"
               f"<table class=\"wrapped\"><tbody><tr style=\"background-color: rgb(244,245,247);\">"
               f"<th><p>Metric</p></th><th><p>Today</p></th><th><p>Prev</p></th><th><p>&Delta; %</p></th></tr>"
               f"<tr><td><p>Transactions</p></td><td><p>{cur['tx']:,}</p></td><td><p>{pre['tx']:,}</p></td><td><p>{_pct(cur['tx'],pre['tx']):+.1f}%</p></td></tr>"
               f"<tr><td><p>Revenue (VND)</p></td><td><p>{cur['rev']:,}</p></td><td><p>{pre['rev']:,}</p></td><td><p>{_pct(cur['rev'],pre['rev']):+.1f}%</p></td></tr>"
               f"<tr><td><p>Success rate</p></td><td><p>{cur['sr']*100:.1f}%</p></td><td><p>{pre['sr']*100:.1f}%</p></td><td><p>{(cur['sr']-pre['sr'])*100:+.2f}pp</p></td></tr>"
               f"</tbody></table><p><strong>By product</strong></p><table class=\"wrapped\"><tbody>"
               f"<tr style=\"background-color: rgb(244,245,247);\"><th><p>Product</p></th><th><p>Transactions</p></th><th><p>Revenue (VND)</p></th><th><p>Success rate</p></th></tr>"
               f"{rows}</tbody></table>")
    new_body = section + body  # newest on top
    payload = json.dumps({"type": "page", "title": d["title"],
                          "version": {"number": ver + 1, "message": f"MBS daily report {k['today']}"},
                          "body": {"storage": {"value": new_body, "representation": "storage"}}}).encode()
    put = urllib.request.Request(base + "?notifyWatchers=false", data=payload, method="PUT", headers=h)
    r = urllib.request.urlopen(put, timeout=30)
    return json.load(r)["version"]["number"]


def post_telegram(text):
    import urllib.request, json
    tok = os.getenv("TELEGRAM_BOT_TOKEN", "").strip()
    chat = os.getenv("TELEGRAM_CHAT_ID", "").strip()
    if not (tok and chat):
        return "skipped (TELEGRAM_BOT_TOKEN/CHAT_ID not set)"
    body = json.dumps({"chat_id": chat, "text": text}).encode()
    req = urllib.request.Request(f"https://api.telegram.org/bot{tok}/sendMessage", data=body,
                                 headers={"Content-Type": "application/json"})
    urllib.request.urlopen(req, timeout=20)
    return "posted"


def main():
    k = compute(load_rows())
    print(report_text(k))
    if "--excel" in sys.argv or "--all" in sys.argv:
        print("\n[excel]", write_excel(k))
    if "--confluence" in sys.argv or "--all" in sys.argv:
        print("[confluence] page version ->", paste_confluence(k))
    if "--telegram" in sys.argv or "--all" in sys.argv:
        print("[telegram]", post_telegram(report_text(k)))


if __name__ == "__main__":
    main()
